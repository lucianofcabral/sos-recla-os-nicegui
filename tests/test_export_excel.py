"""Tests for Excel export of credit notes."""

from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from src.application.export_excel import exportar_notas_credito_excel
from src.domain.dto.read import NotaCreditoSinAsignarItem


def test_exportar_notas_credito_excel_desde_dtos() -> None:
    items = [
        NotaCreditoSinAsignarItem(
            credit_note_id=1,
            pago_id=10,
            monto=5000.0,
            fecha_pago=date(2026, 1, 15),
            dominio='AB123CD',
            cliente='ACME S.A.',
            poliza='P-001',
            nro_gestion=1001,
        ),
        NotaCreditoSinAsignarItem(
            credit_note_id=2,
            pago_id=11,
            monto=12500.5,
            fecha_pago=date(2026, 2, 20),
            dominio='XY987ZT',
            cliente='Juan Perez',
            poliza='P-002',
            nro_gestion=None,
        ),
    ]

    excel_bytes = exportar_notas_credito_excel(items)
    assert isinstance(excel_bytes, bytes)
    assert len(excel_bytes) > 0

    wb = load_workbook(BytesIO(excel_bytes))
    ws = wb.active
    assert ws is not None
    assert ws.title == 'Notas de Crédito'

    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == (
        'Fecha',
        'Dominio',
        'Cliente',
        'Póliza',
        'Nro. Gestión',
        'Importe',
    )
    assert rows[1] == (
        '15/01/2026',
        'AB123CD',
        'ACME S.A.',
        'P-001',
        1001,
        5000.0,
    )
    assert rows[2] == (
        '20/02/2026',
        'XY987ZT',
        'Juan Perez',
        'P-002',
        None,
        12500.5,
    )


def test_exportar_notas_credito_excel_desde_dicts() -> None:
    dict_items = [
        {
            'credit_note_id': 1,
            'fecha': '15/01/2026',
            'dominio': 'AB123CD',
            'cliente': 'ACME S.A.',
            'poliza': 'P-001',
            'nro_gestion': '1001',
            'monto': 5000.0,
        }
    ]

    excel_bytes = exportar_notas_credito_excel(dict_items)
    wb = load_workbook(BytesIO(excel_bytes))
    ws = wb.active
    assert ws is not None
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == (
        'Fecha',
        'Dominio',
        'Cliente',
        'Póliza',
        'Nro. Gestión',
        'Importe',
    )
    assert rows[1] == (
        '15/01/2026',
        'AB123CD',
        'ACME S.A.',
        'P-001',
        1001,
        5000.0,
    )
