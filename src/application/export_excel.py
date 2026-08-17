"""Excel export helpers for domain read models."""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from src.domain.dto.read import NotaCreditoSinAsignarItem


def exportar_notas_credito_excel(
    ncs: list[NotaCreditoSinAsignarItem] | list[dict[str, Any]],
) -> bytes:
    """Generate an Excel .xlsx workbook (as bytes) for the given credit notes."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = 'Notas de Crédito'

    headers = ['Fecha', 'Dominio', 'Cliente', 'Póliza', 'Nro. Gestión', 'Importe']
    ws.append(headers)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(
        start_color='1976D2', end_color='1976D2', fill_type='solid'
    )
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for item in ncs:
        if isinstance(item, NotaCreditoSinAsignarItem):
            fecha_str = (
                item.fecha_pago.strftime('%d/%m/%Y') if item.fecha_pago else None
            )
            dominio = item.dominio or None
            cliente = item.cliente or None
            poliza = item.poliza or None
            nro_gestion = item.nro_gestion if item.nro_gestion is not None else None
            monto = item.monto if item.monto is not None else 0.0
        else:
            fecha_val = item.get('fecha_pago') or item.get('fecha')
            fecha_str = str(fecha_val) if fecha_val and fecha_val != '-' else None
            dominio_val = item.get('dominio')
            dominio = str(dominio_val) if dominio_val and dominio_val != '-' else None
            cliente_val = item.get('cliente')
            cliente = str(cliente_val) if cliente_val and cliente_val != '-' else None
            poliza_val = item.get('poliza')
            poliza = str(poliza_val) if poliza_val and poliza_val != '-' else None
            nro_val = item.get('nro_gestion')
            nro_gestion = (
                int(nro_val)
                if nro_val is not None and str(nro_val).isdigit()
                else (str(nro_val) if nro_val and nro_val != '-' else None)
            )
            monto_val = item.get('monto') or item.get('monto_raw') or 0.0
            monto = float(monto_val)

        ws.append([fecha_str, dominio, cliente, poliza, nro_gestion, monto])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
