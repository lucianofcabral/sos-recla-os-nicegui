"""Excel import for SOS reclamos (upsert keyed by Nro de Gestion).

This module parses a ``Gestion Reclamos Y Reintegros.xlsx`` workbook and upserts
SOS reclamos through a ``UnitOfWorkPort``. Rows are keyed by ``nro_gestion``:
missing rows are created, existing rows only overwrite the fields whose parsed
value is not None (blank text cells never wipe a stored value; an empty Póliza
cell follows the legacy ``''`` convention and does overwrite). Parsing is pure
and side-effect free (``parse_excel_sos``); the import pipeline commits in
batches and supports a ``dry_run`` that only counts.
"""

from __future__ import annotations

import unicodedata
from collections.abc import Iterable, Sequence
from dataclasses import dataclass
from datetime import date, datetime, time
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from src.application.import_gestiones import parse_fecha
from src.domain.domain_enums import TipoReclamoEnum
from src.domain.models.entities import Reclamo, ReclamoSos
from src.domain.ports.unit_of_work import UnitOfWorkPort

BATCH_SIZE = 100

COLUMN_KEYS: dict[str, tuple[str, ...]] = {
    'nro_gestion': ('n° gestión',),
    'fecha': ('fecha',),
    'cliente': ('cliente',),
    'dominio': ('dominio',),
    'poliza': ('póliza',),
    'motivo': ('motivo',),
    'usuario_carga': ('usuario carga',),
    'usuario_respuesta': ('usuario respuesta',),
    'status': ('estado',),
    'itr': ('itr',),
}

BASE_FIELDS: tuple[str, ...] = ('cliente', 'poliza', 'dominio')
SOS_FIELDS: tuple[str, ...] = (
    'motivo',
    'usuario_carga',
    'usuario_respuesta',
    'status',
    'itr',
)


@dataclass(frozen=True)
class SosExcelRow:
    """Normalized SOS reclamo row read from the Excel workbook."""

    fecha: date | None
    nro_gestion: int
    cliente: str | None
    dominio: str | None
    poliza: str
    motivo: str | None
    usuario_carga: str | None
    usuario_respuesta: str | None
    status: str | None
    itr: int | None


@dataclass(frozen=True)
class ImportExcelSosReport:
    """Result of an SOS Excel import (or dry-run preview)."""

    creados: int
    actualizados: int
    errores: list[str]


def _celda(fila: Sequence[object], idx: int | None) -> object:
    if idx is None or idx >= len(fila):
        return None
    return fila[idx]


def _optional_str(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _fila_es_vacia(fila: Sequence[object]) -> bool:
    return all(c is None or (isinstance(c, str) and not c.strip()) for c in fila)


def _parse_nro_gestion(value: object) -> int | None:
    """Parse ``N° Gestión`` to a positive int; None when invalid/blank/zero."""
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, float):
        if value.is_integer() and value > 0:
            return int(value)
        return None
    if isinstance(value, int):
        return value if value > 0 else None
    text = str(value).strip()
    if not text:
        return None
    try:
        parsed = int(text)
    except ValueError:
        return None
    return parsed if parsed > 0 else None


def _parse_itr(value: object) -> int | None:
    """Parse ``ITR``; ``0``/``'0'``/blank map to None (legacy ``itr or None``)."""
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return None if value == 0 else value
    if isinstance(value, float):
        if not value.is_integer():
            return None
        return None if int(value) == 0 else int(value)
    text = str(value).strip()
    if not text or text in {'0', '0.0'}:
        return None
    try:
        itr = int(text)
    except ValueError:
        return None
    return None if itr == 0 else itr


def parse_excel_sos(contenido: bytes) -> tuple[list[SosExcelRow], list[str]]:
    """Parse an SOS xlsx workbook into valid rows plus per-row error messages.

    Raises ``ValueError`` for structural problems: a file that cannot be read,
    a workbook without sheets, or a missing ``N° Gestión`` header.
    """
    try:
        workbook = load_workbook(BytesIO(contenido), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError('El archivo no se pudo leer como Excel') from exc
    try:
        if not workbook.worksheets:
            raise ValueError('El archivo no contiene hojas de cálculo')
        hoja = workbook.worksheets[0]
        # Algunos exportadores declaran un <dimension> incorrecto (p. ej. "A1")
        # y openpyxl solo itera esa celda; recalcular la dimensión real.
        if hasattr(hoja, 'reset_dimensions'):
            hoja.reset_dimensions()
        filas = hoja.iter_rows(values_only=True)
        encabezado = next(filas, None)
        if encabezado is None:
            raise ValueError('El archivo no contiene una fila de encabezado')
        indices = _indices_por_columna(encabezado)
        if 'nro_gestion' not in indices:
            detectadas = ', '.join(
                str(celda) for celda in encabezado if celda is not None
            )
            raise ValueError(
                'Falta la columna "N° Gestión" en el encabezado.'
                + (f' Columnas detectadas: {detectadas}' if detectadas else '')
            )
        return _parse_filas(filas, indices)
    finally:
        workbook.close()


def _indices_por_columna(encabezado: Sequence[object]) -> dict[str, int]:
    indices: dict[str, int] = {}
    for idx, celda in enumerate(encabezado):
        if celda is None or not str(celda).strip():
            continue
        clave = _clave_columna(str(celda))
        if clave is not None and clave not in indices:
            indices[clave] = idx
    return indices


def _normalizar(texto: str) -> str:
    """Normalize a header for tolerant matching (case, accents, N°/Nº/NRO)."""
    texto = unicodedata.normalize('NFKD', texto).casefold()
    return ''.join(c for c in texto if c.isalnum())


def _es_nro_gestion(nombre: str) -> bool:
    """Best-effort detection of the N° Gestión column across exporter variants."""
    norm = _normalizar(nombre)
    if 'gestion' not in norm:
        return False
    return norm.startswith('n') or 'nro' in norm or 'num' in norm


def _clave_columna(nombre: str) -> str | None:
    norm = _normalizar(nombre)
    if _es_nro_gestion(nombre):
        return 'nro_gestion'
    for key, alternativas in COLUMN_KEYS.items():
        if any(norm == _normalizar(alt) for alt in alternativas):
            return key
    return None


def _parse_filas(
    filas: Iterable[Sequence[object]], indices: dict[str, int]
) -> tuple[list[SosExcelRow], list[str]]:
    filas_validas: list[SosExcelRow] = []
    errores: list[str] = []
    for idx, fila in enumerate(filas):
        if _fila_es_vacia(fila):
            continue
        nro_fila = idx + 2
        nro_gestion = _parse_nro_gestion(_celda(fila, indices.get('nro_gestion')))
        if nro_gestion is None:
            errores.append(f'Fila {nro_fila}: N° Gestión inválido')
            continue
        filas_validas.append(
            SosExcelRow(
                fecha=parse_fecha(_celda(fila, indices.get('fecha'))),
                nro_gestion=nro_gestion,
                cliente=_optional_str(_celda(fila, indices.get('cliente'))),
                dominio=_optional_str(_celda(fila, indices.get('dominio'))),
                poliza=_optional_str(_celda(fila, indices.get('poliza'))) or '',
                motivo=_optional_str(_celda(fila, indices.get('motivo'))),
                usuario_carga=_optional_str(_celda(fila, indices.get('usuario_carga'))),
                usuario_respuesta=_optional_str(
                    _celda(fila, indices.get('usuario_respuesta'))
                ),
                status=_optional_str(_celda(fila, indices.get('status'))),
                itr=_parse_itr(_celda(fila, indices.get('itr'))),
            )
        )
    return filas_validas, errores


def importar_excel_sos(
    *,
    contenido: bytes,
    uow: UnitOfWorkPort,
    dry_run: bool = False,
) -> ImportExcelSosReport:
    """Upsert SOS reclamos from an Excel workbook, keyed by ``nro_gestion``.

    Returns a report with created/updated counts and the collected row errors.
    ``dry_run=True`` only counts and performs existence lookups; nothing is
    written or committed.
    """
    filas, errores = parse_excel_sos(contenido)
    creados = 0
    actualizados = 0
    pendientes = 0
    for fila in filas:
        existe = uow.reclamos_sos.get_by_nro_gestion(fila.nro_gestion) is not None
        if existe:
            actualizados += 1
        else:
            creados += 1
        if dry_run:
            continue
        if existe:
            _actualizar_sos(uow, fila)
        else:
            _crear_sos(uow, fila)
        pendientes += 1
        if pendientes >= BATCH_SIZE:
            uow.commit()
            pendientes = 0
    if not dry_run and pendientes:
        uow.commit()
    return ImportExcelSosReport(
        creados=creados, actualizados=actualizados, errores=errores
    )


def _created_at(fecha: date | None) -> datetime:
    if fecha is not None:
        return datetime.combine(fecha, time.min)
    return datetime.now()


def _crear_sos(uow: UnitOfWorkPort, fila: SosExcelRow) -> None:
    creado = _created_at(fila.fecha)
    reclamo = uow.reclamos.save(
        Reclamo(
            tipo_reclamo=TipoReclamoEnum.SOS,
            cliente=fila.cliente,
            dominio=fila.dominio,
            poliza=fila.poliza,
            importe_reclamado=0.0,
            active=True,
            created_at=creado,
            updated_at=creado,
        )
    )
    assert reclamo.id is not None
    uow.reclamos_sos.save(
        ReclamoSos(
            reclamo_id=reclamo.id,
            reclamo=reclamo,
            nro_gestion=fila.nro_gestion,
            motivo=fila.motivo,
            usuario_carga=fila.usuario_carga,
            usuario_respuesta=fila.usuario_respuesta,
            status=fila.status,
            itr=fila.itr,
        )
    )


def _cambios(fila: SosExcelRow) -> tuple[dict[str, Any], dict[str, Any]]:
    base: dict[str, Any] = {
        f: getattr(fila, f) for f in BASE_FIELDS if getattr(fila, f) is not None
    }
    sos: dict[str, Any] = {
        f: getattr(fila, f) for f in SOS_FIELDS if getattr(fila, f) is not None
    }
    return base, sos


def _actualizar_sos(uow: UnitOfWorkPort, fila: SosExcelRow) -> None:
    sos = uow.reclamos_sos.get_by_nro_gestion(fila.nro_gestion)
    assert sos is not None
    reclamo = sos.reclamo
    if reclamo is None:
        assert sos.reclamo_id is not None
        reclamo = uow.reclamos.get(sos.reclamo_id)
    base, sos_cambios = _cambios(fila)
    if base:
        reclamo = uow.reclamos.update(
            Reclamo.model_validate({**reclamo.model_dump(), **base})
        )
    sos = ReclamoSos.model_validate(
        {**sos.model_dump(), **sos_cambios, 'reclamo': reclamo}
    )
    uow.reclamos_sos.update(sos)
